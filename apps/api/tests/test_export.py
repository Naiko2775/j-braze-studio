"""Tests pour l'export Data Model (Excel et CSV)."""
import io
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from fastapi.testclient import TestClient

from main import app
from services.data_model.demo_data import DEMO_RESULTS
from services.data_model.exporter import export_to_excel, export_to_csv

client = TestClient(app)


# ── Tests unitaires export Excel ────────────────────────────────────────────


def test_export_excel_returns_bytes():
    """L'export Excel retourne des bytes non vides."""
    data = export_to_excel(DEMO_RESULTS)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_export_excel_has_correct_sheets():
    """Le fichier Excel contient les 6 onglets attendus."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    expected_sheets = {"Synthese", "Custom Attributes", "Custom Events", "Segments", "Messaging", "Hierarchie"}
    assert set(wb.sheetnames) == expected_sheets


def test_export_excel_synthese_has_banner():
    """L'onglet Synthese contient le bandeau JAKALA."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Synthese"]
    assert "JAKALA" in str(ws.cell(row=1, column=1).value)


def test_export_excel_custom_attributes_has_data():
    """L'onglet Custom Attributes contient des lignes de donnees."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Custom Attributes"]
    # Row 3 = header, row 4+ = data
    assert ws.cell(row=3, column=1).value == "Nom"
    assert ws.cell(row=4, column=1).value is not None  # au moins une ligne de donnees


def test_export_excel_custom_events_has_headers():
    """L'onglet Custom Events a les bons headers."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Custom Events"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, 5)]
    assert headers == ["Nom", "Description", "Proprietes", "Declencheur"]


def test_export_excel_segments_has_data():
    """L'onglet Segments contient des donnees."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Segments"]
    # Il y a au moins un segment dans les donnees demo
    assert ws.cell(row=4, column=1).value is not None


def test_export_excel_hierarchy_has_data():
    """L'onglet Hierarchie contient des lignes."""
    data = export_to_excel(DEMO_RESULTS)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Hierarchie"]
    assert ws.cell(row=4, column=1).value is not None


# ── Tests unitaires export CSV ──────────────────────────────────────────────


def test_export_csv_returns_string():
    """L'export CSV retourne une chaine non vide."""
    data = export_to_csv(DEMO_RESULTS)
    assert isinstance(data, str)
    assert len(data) > 0


def test_export_csv_uses_semicolon():
    """Le CSV utilise le point-virgule comme separateur."""
    data = export_to_csv(DEMO_RESULTS)
    lines = data.strip().split("\n")
    # Les headers de section ont au moins un ;
    header_lines = [l for l in lines if "Nom" in l]
    assert any(";" in l for l in header_lines)


def test_export_csv_has_all_sections():
    """Le CSV contient toutes les sections."""
    data = export_to_csv(DEMO_RESULTS)
    assert "=== Custom Attributes ===" in data
    assert "=== Custom Events ===" in data
    assert "=== Segments ===" in data
    assert "=== Messaging ===" in data
    assert "=== Hierarchie ===" in data


# ── Tests endpoint ──────────────────────────────────────────────────────────


def test_export_endpoint_with_existing_analysis():
    """L'endpoint export fonctionne avec une analyse existante (mode demo)."""
    # D'abord creer une analyse
    resp = client.post(
        "/api/data-model/analyze",
        json={"use_cases": ["Campagne test"], "demo": True},
    )
    assert resp.status_code == 200
    analysis_id = resp.json().get("id")
    assert analysis_id is not None

    # Exporter en Excel
    resp_excel = client.get(f"/api/data-model/export/{analysis_id}?format=excel")
    assert resp_excel.status_code == 200
    assert "spreadsheetml" in resp_excel.headers["content-type"]
    assert len(resp_excel.content) > 0

    # Exporter en CSV
    resp_csv = client.get(f"/api/data-model/export/{analysis_id}?format=csv")
    assert resp_csv.status_code == 200
    assert "csv" in resp_csv.headers["content-type"]
    assert len(resp_csv.content) > 0


def test_export_endpoint_not_found():
    """L'endpoint retourne 404 pour un ID inexistant."""
    resp = client.get("/api/data-model/export/inexistant-id-000?format=excel")
    assert resp.status_code == 404


def test_export_endpoint_invalid_format():
    """L'endpoint rejette un format invalide."""
    resp = client.get("/api/data-model/export/some-id?format=pdf")
    assert resp.status_code == 422
