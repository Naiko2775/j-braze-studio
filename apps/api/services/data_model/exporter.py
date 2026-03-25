"""Export des resultats d'analyse Data Model en Excel et CSV."""
import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────────

NAVY = "040066"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5FA"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color=NAVY),
    right=Side(style="thin", color="CCCCCC"),
)

CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=11)
BANNER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=13)
BANNER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


# ── Helpers ─────────────────────────────────────────────────────────────────


def _add_banner(ws, text: str, col_count: int = 6):
    """Ajoute un bandeau JAKALA en haut de l'onglet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = BANNER_FONT
    cell.fill = BANNER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36


def _write_header_row(ws, row: int, headers: list[str]):
    """Ecrit une ligne d'en-tete stylisee."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    ws.row_dimensions[row].height = 28


def _write_data_row(ws, row: int, values: list, alt: bool = False):
    """Ecrit une ligne de donnees."""
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGNMENT
        if alt:
            cell.fill = ALT_ROW_FILL


def _auto_column_width(ws, min_width: int = 12, max_width: int = 55):
    """Ajuste la largeur des colonnes au contenu."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, longest)
        width = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# ── Extraction des donnees ──────────────────────────────────────────────────


def _extract_custom_attributes(analysis_result: dict) -> list[list[str]]:
    """Extrait les custom attributes de tous les use cases."""
    rows = []
    for uc in analysis_result.get("use_case_analysis", []):
        uc_name = uc.get("use_case", "")
        for data in uc.get("required_data", []):
            entity = data.get("entity", "")
            for cf in data.get("custom_fields", []):
                rows.append([
                    cf.get("name", ""),
                    cf.get("type", ""),
                    cf.get("description", ""),
                    entity,
                    "Oui" if entity == "User Profile" else "Non",
                ])
    # Deduplicate by name (keep first)
    seen = set()
    unique = []
    for row in rows:
        if row[0] not in seen:
            seen.add(row[0])
            unique.append(row)
    return unique


def _extract_custom_events(analysis_result: dict) -> list[list[str]]:
    """Extrait les custom events."""
    rows = []
    for uc in analysis_result.get("use_case_analysis", []):
        for data in uc.get("required_data", []):
            if data.get("entity") != "Custom Events":
                continue
            trigger = uc.get("messaging", {}).get("trigger_details", "")
            for cf in data.get("custom_fields", []):
                # Extraire les proprietes depuis la description
                desc = cf.get("description", "")
                props = ""
                if "proprietes:" in desc.lower() or "proprietes :" in desc.lower():
                    parts = desc.lower().split("proprietes")
                    if len(parts) > 1:
                        props = parts[1].lstrip(": ").strip()
                rows.append([
                    cf.get("name", ""),
                    desc,
                    props,
                    trigger,
                ])
    seen = set()
    unique = []
    for row in rows:
        if row[0] not in seen:
            seen.add(row[0])
            unique.append(row)
    return unique


def _extract_segments(analysis_result: dict) -> list[list[str]]:
    """Extrait les segments."""
    rows = []
    for uc in analysis_result.get("use_case_analysis", []):
        for seg in uc.get("segments", []):
            filters = "\n".join(seg.get("filters", []))
            rows.append([
                seg.get("name", ""),
                seg.get("description", f"Segment pour: {uc.get('use_case', '')}"),
                filters,
                "",  # Taille estimee -- pas dans les donnees
            ])
    return rows


def _extract_messaging(analysis_result: dict) -> list[list[str]]:
    """Extrait les configurations messaging."""
    rows = []
    for uc in analysis_result.get("use_case_analysis", []):
        messaging = uc.get("messaging")
        if not messaging:
            continue
        channels = messaging.get("channels", [])
        trigger_type = messaging.get("trigger_type", "")
        trigger_details = messaging.get("trigger_details", "")
        for ch in channels:
            rows.append([
                ch,
                trigger_type,
                trigger_details,
                "",  # Frequence
                uc.get("use_case", ""),
            ])
    return rows


def _extract_hierarchy_rows(
    nodes: list[dict], rows: list[list[str]] | None = None, parent: str = "", level: int = 0
) -> list[list[str]]:
    """Aplatit l'arbre hierarchique en lignes."""
    if rows is None:
        rows = []
    for node in nodes:
        entity = node.get("entity", node.get("name", ""))
        attrs = ", ".join(node.get("attributes_used", []))
        indent = "  " * level
        rows.append([
            f"{indent}{entity}",
            parent,
            attrs,
            str(level),
        ])
        children = node.get("children", [])
        if children:
            _extract_hierarchy_rows(children, rows, parent=entity, level=level + 1)
    return rows


# ── Export Excel ────────────────────────────────────────────────────────────


def export_to_excel(analysis_result: dict) -> bytes:
    """Genere un fichier Excel professionnel a partir du resultat d'analyse."""
    wb = Workbook()

    # ── Onglet Synthese ─────────────────────────────────────────────────
    ws_synth = wb.active
    ws_synth.title = "Synthese"
    _add_banner(ws_synth, "JAKALA - J-Braze Studio", col_count=3)

    ws_synth.cell(row=3, column=1, value="Data Model - Rapport d'analyse").font = TITLE_FONT
    ws_synth.cell(row=4, column=1, value=f"Date : {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = CELL_FONT

    analyses = analysis_result.get("use_case_analysis", [])
    ws_synth.cell(row=6, column=1, value="Use cases analyses :").font = SUBTITLE_FONT
    for i, uc in enumerate(analyses):
        ws_synth.cell(row=7 + i, column=1, value=f"{i + 1}. {uc.get('use_case', '')}").font = CELL_FONT
        ws_synth.cell(row=7 + i, column=2, value=uc.get("description", "")).font = CELL_FONT

    summary_row = 7 + len(analyses) + 1
    ws_synth.cell(row=summary_row, column=1, value="Resume").font = SUBTITLE_FONT

    # Count entities
    entities_set = set()
    custom_attrs = set()
    segments_count = 0
    channels_set = set()
    for uc in analyses:
        for d in uc.get("required_data", []):
            entities_set.add(d.get("entity", ""))
            for cf in d.get("custom_fields", []):
                custom_attrs.add(cf.get("name", ""))
        segments_count += len(uc.get("segments", []))
        for ch in uc.get("messaging", {}).get("channels", []):
            channels_set.add(ch)

    stats = [
        ("Entites Braze identifiees", str(len(entities_set))),
        ("Custom Attributes / Events", str(len(custom_attrs))),
        ("Segments a creer", str(segments_count)),
        ("Canaux de messaging", ", ".join(sorted(channels_set))),
    ]
    for i, (label, value) in enumerate(stats):
        ws_synth.cell(row=summary_row + 1 + i, column=1, value=label).font = CELL_FONT
        ws_synth.cell(row=summary_row + 1 + i, column=2, value=value).font = Font(name="Calibri", bold=True, size=10)

    _auto_column_width(ws_synth)

    # ── Onglet Custom Attributes ────────────────────────────────────────
    ws_attrs = wb.create_sheet("Custom Attributes")
    _add_banner(ws_attrs, "JAKALA - J-Braze Studio", col_count=5)
    headers_attrs = ["Nom", "Type", "Description", "Source", "Obligatoire"]
    _write_header_row(ws_attrs, 3, headers_attrs)
    attr_rows = _extract_custom_attributes(analysis_result)
    for i, row in enumerate(attr_rows):
        _write_data_row(ws_attrs, 4 + i, row, alt=(i % 2 == 1))
    _auto_column_width(ws_attrs)

    # ── Onglet Custom Events ────────────────────────────────────────────
    ws_events = wb.create_sheet("Custom Events")
    _add_banner(ws_events, "JAKALA - J-Braze Studio", col_count=4)
    headers_events = ["Nom", "Description", "Proprietes", "Declencheur"]
    _write_header_row(ws_events, 3, headers_events)
    event_rows = _extract_custom_events(analysis_result)
    for i, row in enumerate(event_rows):
        _write_data_row(ws_events, 4 + i, row, alt=(i % 2 == 1))
    _auto_column_width(ws_events)

    # ── Onglet Segments ─────────────────────────────────────────────────
    ws_seg = wb.create_sheet("Segments")
    _add_banner(ws_seg, "JAKALA - J-Braze Studio", col_count=4)
    headers_seg = ["Nom", "Description", "Filtres", "Taille estimee"]
    _write_header_row(ws_seg, 3, headers_seg)
    seg_rows = _extract_segments(analysis_result)
    for i, row in enumerate(seg_rows):
        _write_data_row(ws_seg, 4 + i, row, alt=(i % 2 == 1))
    _auto_column_width(ws_seg)

    # ── Onglet Messaging ────────────────────────────────────────────────
    ws_msg = wb.create_sheet("Messaging")
    _add_banner(ws_msg, "JAKALA - J-Braze Studio", col_count=5)
    headers_msg = ["Canal", "Type", "Declencheur", "Frequence", "Contenu"]
    _write_header_row(ws_msg, 3, headers_msg)
    msg_rows = _extract_messaging(analysis_result)
    for i, row in enumerate(msg_rows):
        _write_data_row(ws_msg, 4 + i, row, alt=(i % 2 == 1))
    _auto_column_width(ws_msg)

    # ── Onglet Hierarchie ───────────────────────────────────────────────
    ws_hier = wb.create_sheet("Hierarchie")
    _add_banner(ws_hier, "JAKALA - J-Braze Studio", col_count=4)
    headers_hier = ["Entite", "Parent", "Attributs utilises", "Niveau"]
    _write_header_row(ws_hier, 3, headers_hier)
    hierarchy = analysis_result.get("data_hierarchy", [])
    hier_rows = _extract_hierarchy_rows(hierarchy)
    for i, row in enumerate(hier_rows):
        _write_data_row(ws_hier, 4 + i, row, alt=(i % 2 == 1))
    _auto_column_width(ws_hier)

    # ── Ecriture en bytes ───────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── Export CSV ──────────────────────────────────────────────────────────────


def export_to_csv(analysis_result: dict) -> str:
    """Exporte toutes les donnees en CSV flat (separateur point-virgule)."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    # ── Section Custom Attributes
    writer.writerow(["=== Custom Attributes ==="])
    writer.writerow(["Nom", "Type", "Description", "Source", "Obligatoire"])
    for row in _extract_custom_attributes(analysis_result):
        writer.writerow(row)
    writer.writerow([])

    # ── Section Custom Events
    writer.writerow(["=== Custom Events ==="])
    writer.writerow(["Nom", "Description", "Proprietes", "Declencheur"])
    for row in _extract_custom_events(analysis_result):
        writer.writerow(row)
    writer.writerow([])

    # ── Section Segments
    writer.writerow(["=== Segments ==="])
    writer.writerow(["Nom", "Description", "Filtres", "Taille estimee"])
    for row in _extract_segments(analysis_result):
        writer.writerow(row)
    writer.writerow([])

    # ── Section Messaging
    writer.writerow(["=== Messaging ==="])
    writer.writerow(["Canal", "Type", "Declencheur", "Frequence", "Contenu"])
    for row in _extract_messaging(analysis_result):
        writer.writerow(row)
    writer.writerow([])

    # ── Section Hierarchie
    writer.writerow(["=== Hierarchie ==="])
    writer.writerow(["Entite", "Parent", "Attributs utilises", "Niveau"])
    hierarchy = analysis_result.get("data_hierarchy", [])
    for row in _extract_hierarchy_rows(hierarchy):
        writer.writerow(row)

    return output.getvalue()
